from typing import List, Any, Union

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook


def create_column():
    url = "https://finance.naver.com/item/main.nhn?code=005930" # 삼성전자
    res = requests.get(url)
    html = BeautifulSoup(res.text, "html.parser")

    cop_anal = html.find("div", {"class", "section cop_analysis"})
    sub_section_div = cop_anal.find("div", {"class", "sub_section"})
    ifrs_table = sub_section_div.find("table", {"class", "tb_type1 tb_num tb_type1_ifrs"})
    
    # 연도정보
    thead = ifrs_table.find("thead")
    all_thead_tr = thead.find_all("tr")
    all_thead_tr_th = all_thead_tr[1].find_all("th")

    # 컬럼 구성
    columns = []
    columns.append("기업명")
    columns.append("시장")

    columns.append("연간 영업이익: " + all_thead_tr_th[2].get_text().strip())
    columns.append(all_thead_tr_th[3].get_text().strip())
    columns.append("YoY(%)")

    columns.append("분기 매출 " + all_thead_tr_th[8].get_text().strip())
    columns.append(all_thead_tr_th[9].get_text().strip())
    columns.append("YoY(%)")
    columns.append("QoQ(%)")

    columns.append("분기 영업이익" + all_thead_tr_th[8].get_text().strip())
    columns.append(all_thead_tr_th[9].get_text().strip())
    columns.append("YoY(%)")
    columns.append("QoQ(%)")

    columns.append("시가총액(억)")
    columns.append("목표시총(억)")
    columns.append("상승여력(%)")
    columns.append("배당율(%)")
    columns.append("멀티플")
    columns.append("업종")
    columns.append("기업개요")

    return columns


def get_multiple_value(name, category):
    name_multiple_tb = {
        "코오롱글로벌": 5,
        "뷰웍스": 15,
        "스튜디오드래곤": 15,
        "에코마케팅": 15,
        "카카오": 30,
        "NAVER": 30,
        "에코프로비엠": 20,
        "엘앤에프": 20,
        "포스코케미칼": 20,
        "한컴위드": 15,
        "한글과컴퓨터": 15,
    }
    category_multiple_tb = {
        "조선": 6,
        "증권": 5,
        "은행": 5,
        "해운사": 7,
        "건설": 8,
        "호텔,레스토랑,레저": 7,
        "IT서비스": 10,
        "양방향미디어와서비스": 15,
        "통신장비": 10,
        "게임엔터테인먼트": 15,
        "건강관리장비와용품": 15,
        "소프트웨어": 15,
        "제약": 20,
    }

    korean_multiple = 10
    if name in name_multiple_tb:
        return name_multiple_tb.get(name)
    if category in category_multiple_tb:
        return category_multiple_tb.get(category)
    return korean_multiple


columns = create_column()
val_result_wb = Workbook()
val_result_ws = val_result_wb.active
val_result_ws.append(columns)

# 실제 내용은 html이기 때문에 read_html로 읽는다.
# 종목코드를 빈자리는 0으로 채워진 6자리 문자열로 변환한다.
stock_df = pd.read_html('http://kind.krx.co.kr/corpgeneral/corpList.do?method=download&searchType=13', header=0)[0]
stock_df['종목코드'] = stock_df['종목코드'].map(lambda x: f'{x:0>6}')
stock_arr = stock_df.to_numpy()

for stock_id in range(0, len(stock_arr)):
    val_result_ws.append([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])  # 초기값
    val_result_ws.cell(stock_id + 2, 1, stock_arr[stock_id][0])  # 기업명

    url = "https://finance.naver.com/item/main.nhn?code=" + stock_arr[stock_id][1]
    res = requests.get(url)
    html = BeautifulSoup(res.text, "html.parser")

    # kospi or kosdaq
    stock_type = "코넥스"
    new_totalinfo_div = html.find("div", {"class", "new_totalinfo"})
    if not new_totalinfo_div:
        continue
    dl_totalinfo = new_totalinfo_div.find("dl", {"class", "blind"})
    dd_totalinfo_all = dl_totalinfo.find_all("dd")
    dd_text = dd_totalinfo_all[2].get_text().strip()
    if "코스닥" in dd_text:
        stock_type = "코스닥"
    elif "코스피" in dd_text:
        stock_type = "코스피"

    cop_anal = html.find("div", {"class", "section cop_analysis"})
    if not cop_anal:
        continue
    sub_section_div = cop_anal.find("div", {"class", "sub_section"})
    ifrs_table = sub_section_div.find(
        "table", {"class", "tb_type1 tb_num tb_type1_ifrs"})
    tbody = ifrs_table.find("tbody")
    all_tr = tbody.find_all("tr")
    if len(all_tr) < 14:
        continue

    # 4년 매출
    sales = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    sales_all_td = all_tr[0].find_all("td")
    for i in range(0, len(sales)):
        sales_text = sales_all_td[i].get_text().strip().replace(",", "")
        if sales_text:
            if sales_text[0] == '-':
                if len(sales_text) > 1:
                    sales_text = sales_text[1:]
                    sales[i] = int(sales_text) * -1
            else:
                sales[i] = int(sales_text)

    # 4년 영업이익, 6분기 영업이익
    profits = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    profits_all_td = all_tr[1].find_all("td")
    for i in range(0, len(profits)):
        profit_text = profits_all_td[i].get_text().strip().replace(",", "")
        if profit_text:
            if profit_text[0] == '-':
                if len(profit_text) > 1:
                    profit_text = profit_text[1:]
                    profits[i] = int(profit_text) * -1
            else:
                profits[i] = int(profit_text)

    # 배당률
    dividend_rate = 0.0
    dividend_rate_td = all_tr[14].find_all("td")
    dividend_rate_text = dividend_rate_td[3].get_text().strip()
    if dividend_rate_text and dividend_rate_text[0] != '-':
        dividend_rate = float(dividend_rate_text)
    elif dividend_rate_td[2].get_text().strip():
        dividend_rate_text = dividend_rate_td[2].get_text().strip()
        if dividend_rate_text[0] != '-':
            dividend_rate = float(dividend_rate_text)

    # 시총(억)
    market_cap = 0
    cur_price = 0
    trade_compare_div = html.find("div", {"class", "section trade_compare"})
    if trade_compare_div:  # 이미 계산된 값이 있다면 사용
        compare_table = trade_compare_div.find(
            "table", {"class", "tb_type1 tb_num"})
        tbody = compare_table.find("tbody")
        all_tr = tbody.find_all("tr")
        cur_price_text = all_tr[0].find("td").get_text().strip().replace(",", "")
        if cur_price_text:  # 현재가
            cur_price = int(cur_price_text)
        if len(all_tr) > 3:
            market_cap_text = all_tr[3].find("td").get_text().strip().replace(",", "")
            if market_cap_text:
                market_cap = int(market_cap_text)
    if market_cap == 0:
        tab_con1_div = html.find("div", {"class", "tab_con1"})
        stock_total_table = tab_con1_div.find("table")
        stock_total_tr = stock_total_table.find_all("tr")
        stock_total_text = stock_total_tr[2].find("td").get_text().strip().replace(",", "")
        if stock_total_text:
            stock_total = int(stock_total_text)
            market_cap = round(cur_price * stock_total / 100000000)  # 억 단위로 변환
    if market_cap == 0:
        continue

    # 업종
    business_category = stock_arr[stock_id][2]
    trade_compare = html.find("div", {"class", "section trade_compare"})
    if trade_compare:
        trade_compare = trade_compare.find("h4", {"class", "h_sub sub_tit7"})
        trade_compare = trade_compare.find("a")
        if trade_compare.get_text().strip():
            business_category = trade_compare.get_text().strip()

    # 종목 또는 업종에 따른 멀티플
    multiple = get_multiple_value(stock_arr[stock_id][0], business_category)

    # 예상시총: 당해년도 예상 영업이익 -> 다음 분기 예상 영업이익 -> 직전 두 분기 * 2 -> 직전년도 영업이익
    base_val = profits[2]
    if profits[3] > 0:
        base_val = profits[3]
    elif profits[9] > 0:
        base_val = profits[7] + profits[8] + (profits[9] * 2)
    elif profits[7] > 0 and profits[8] > 0:
        base_val = (profits[7] + profits[8]) * 2
    expected_market_cap = base_val * multiple

    # 상승여력(%)
    valuation = round((int(expected_market_cap) / int(market_cap) - 1.0) * 100)
    if valuation < 0:
        valuation = 0

    year_profits_yoy = 0
    if profits[3] > 0 and profits[2] > 0:
        year_profits_yoy = round((profits[3] / profits[2] - 1.0) * 100)
    elif profits[2] < 0 and profits[3] > 0:
        year_profits_yoy = "흑전"
    elif profits[2] < 0 and profits[3] < 0:
        year_profits_yoy = "적지"

    quarter_sales_qoq = 0
    if sales[8] > 0 and sales[9] > 0:
        quarter_sales_qoq = round((sales[9] / sales[8] - 1.0) * 100)
    elif sales[8] < 0 and sales[9] > 0:
        quarter_sales_qoq = "흑전"
    elif sales[8] < 0 and sales[9] < 0:
        quarter_sales_qoq = "적지"

    quarter_sales_yoy = 0
    if sales[5] > 0 and sales[9] > 0:
        quarter_sales_yoy = round((sales[9] / sales[5] - 1.0) * 100)

    quarter_profits_qoq = 0
    if profits[8] > 0 and profits[9] > 0:
        quarter_profits_qoq = round((profits[9] / profits[8] - 1.0) * 100)
    elif profits[8] < 0 and profits[9] > 0:
        quarter_profits_qoq = "흑전"
    elif profits[8] < 0 and profits[9] < 0:
        quarter_profits_qoq = "적지"

    quarter_profits_yoy = 0
    if profits[5] > 0 and profits[9] > 0:
        quarter_profits_yoy = round((profits[9] / profits[5] - 1.0) * 100)

    # 열 추가
    col_data = (stock_type,  # 코스피, 코스닥
                profits[2],  # 영업이익 직전 2년
                profits[3],
                year_profits_yoy,
                sales[8],  # 직전분기 매출
                sales[9],  # 이번분기 매출
                quarter_sales_yoy,  # 전년 동 분기
                quarter_sales_qoq,  # 직전 분기 매출
                profits[8],  # 직전 영업이익
                profits[9],  # 이번 영업이익
                quarter_profits_yoy,  # 전년 동 분기
                quarter_profits_qoq,  # 직전 분기
                market_cap,  # 현시가총액
                expected_market_cap,  # 목표시가총액
                valuation,  # 상승여력
                dividend_rate,  # 배당률
                multiple,  # 멀티플
                business_category,  # 업종
                stock_arr[stock_id][3]  # 기업설명
                )
    for idx in range(2, 21):
        val_result_ws.cell(stock_id + 2, idx, col_data[idx - 2])
    print("#" + str(stock_id) + ": " + stock_arr[stock_id][0])

val_result_wb.save("분기예상실적기준_평가.xlsx")
print("Finished!!")
